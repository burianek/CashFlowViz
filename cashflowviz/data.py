"""Parse the CashFlowViz input workbook and compute a day-by-day balance series."""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass

import openpyxl

STATUS_LABEL = "status"


@dataclass(frozen=True)
class Entry:
    asset: str
    date: dt.date
    amount: float
    is_status: bool


@dataclass(frozen=True)
class DayBalance:
    date: dt.date
    balance: float
    events: tuple[Entry, ...]


def _to_date(value) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise ValueError(f"Cell does not contain a date: {value!r}")


def read_entries(xlsx_path: str) -> list[Entry]:
    """Read (Assets, Date, IN/OUT) rows from the first sheet.

    The header row is located by name so column order/position doesn't matter.
    A row whose Assets value is "STATUS" (case-insensitive) sets the balance
    to an absolute value on that date instead of adding a delta.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row_idx = None
    col_of = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        names = {str(c.value).strip().lower(): c.column for c in row if c.value is not None}
        if {"assets", "date", "in/out"} <= names.keys():
            header_row_idx = row[0].row
            col_of = names
            break
    if header_row_idx is None:
        raise ValueError(
            "Could not find header row with 'Assets', 'Date', 'IN/OUT' columns"
        )

    asset_col = col_of["assets"]
    date_col = col_of["date"]
    amount_col = col_of["in/out"]

    entries: list[Entry] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        asset_cell = ws.cell(row=row[0].row, column=asset_col).value
        date_cell = ws.cell(row=row[0].row, column=date_col).value
        amount_cell = ws.cell(row=row[0].row, column=amount_col).value
        if asset_cell is None and date_cell is None and amount_cell is None:
            continue
        if date_cell is None or amount_cell is None:
            continue
        asset = str(asset_cell).strip() if asset_cell is not None else ""
        date = _to_date(date_cell)
        amount = float(amount_cell)
        is_status = asset.lower() == STATUS_LABEL
        entries.append(Entry(asset=asset, date=date, amount=amount, is_status=is_status))

    if not entries:
        raise ValueError("No data rows found in the workbook")

    return entries


def compute_daily_balance(entries: list[Entry]) -> list[DayBalance]:
    """Expand entries into a continuous day-by-day balance series.

    STATUS rows set the balance to an absolute value on their date. All other
    rows add their amount to the running balance on their date. Days without
    events carry the previous day's balance forward.
    """
    start = min(e.date for e in entries)
    end = max(e.date for e in entries)

    by_date: dict[dt.date, list[Entry]] = {}
    for e in entries:
        by_date.setdefault(e.date, []).append(e)

    days: list[DayBalance] = []
    balance = 0.0
    current = start
    while current <= end:
        todays_events = tuple(by_date.get(current, ()))
        for event in todays_events:
            if event.is_status:
                balance = event.amount
            else:
                balance += event.amount
        days.append(DayBalance(date=current, balance=balance, events=todays_events))
        current += dt.timedelta(days=1)

    return days
