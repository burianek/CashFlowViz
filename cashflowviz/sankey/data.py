"""Parse the FROM/AMOUNT/DATE cashflow format and build a month-hub Sankey graph.

Each row is either income (positive AMOUNT) or an expense (negative AMOUNT),
dated on the day it hits the balance. Rows are grouped by the calendar month
of their DATE into a chronological chain of month "hub" nodes: each hub
receives the balance carried over from the previous month plus that month's
income, and sends out that month's expenses plus whatever balance carries
forward to the next month.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass

import openpyxl

CZECH_MONTHS = {
    # Uppercase, diacritics stripped — matches the "SRPEN" / "ZARI" convention
    # from the reference template rather than full Czech orthography.
    1: "LEDEN",
    2: "UNOR",
    3: "BREZEN",
    4: "DUBEN",
    5: "KVETEN",
    6: "CERVEN",
    7: "CERVENEC",
    8: "SRPEN",
    9: "ZARI",
    10: "RIJEN",
    11: "LISTOPAD",
    12: "PROSINEC",
}


def month_label(year: int, month: int) -> str:
    return f"{CZECH_MONTHS[month]} {year}"


@dataclass(frozen=True)
class FlowEntry:
    name: str
    amount: float
    date: dt.date


@dataclass
class Node:
    id: str
    label: str
    kind: str  # 'source' | 'sink' | 'hub' | 'terminal'
    column: int
    value: float = 0.0
    date: dt.date | None = None
    negative: bool = False  # hub/terminal only: True if it's running a deficit


@dataclass
class Link:
    source: str
    target: str
    value: float
    kind: str  # 'income' | 'expense' | 'carry_pos' | 'carry_neg'


@dataclass
class SankeyGraph:
    nodes: list[Node]
    links: list[Link]


def _to_date(value) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise ValueError(f"Cell does not contain a date: {value!r}")


def read_flow_entries(xlsx_path: str) -> list[FlowEntry]:
    """Read (FROM, AMOUNT, DATE) rows from the first sheet.

    The header row is located by name so column order/position doesn't
    matter. Entries are returned sorted chronologically.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row_idx = None
    col_of = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        names = {str(c.value).strip().lower(): c.column for c in row if c.value is not None}
        if {"from", "amount", "date"} <= names.keys():
            header_row_idx = row[0].row
            col_of = names
            break
    if header_row_idx is None:
        raise ValueError("Could not find header row with 'FROM', 'AMOUNT', 'DATE' columns")

    name_col = col_of["from"]
    amount_col = col_of["amount"]
    date_col = col_of["date"]

    entries: list[FlowEntry] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        name_cell = ws.cell(row=row[0].row, column=name_col).value
        amount_cell = ws.cell(row=row[0].row, column=amount_col).value
        date_cell = ws.cell(row=row[0].row, column=date_col).value
        if name_cell is None and amount_cell is None and date_cell is None:
            continue
        if name_cell is None or amount_cell is None or date_cell is None:
            continue
        entries.append(
            FlowEntry(name=str(name_cell).strip(), amount=float(amount_cell), date=_to_date(date_cell))
        )

    if not entries:
        raise ValueError("No data rows found in the workbook")

    entries.sort(key=lambda e: e.date)
    return entries


def build_graph(entries: list[FlowEntry]) -> SankeyGraph:
    months = sorted({(e.date.year, e.date.month) for e in entries})
    month_index = {ym: i for i, ym in enumerate(months)}

    nodes: dict[str, Node] = {}
    links: list[Link] = []

    def unique_id(base: str, ym: tuple[int, int]) -> str:
        if base not in nodes:
            return base
        candidate = f"{base} ({month_label(*ym)})"
        n = 2
        while candidate in nodes:
            candidate = f"{base} ({month_label(*ym)} #{n})"
            n += 1
        return candidate

    hub_id_of: dict[tuple[int, int], str] = {}
    for ym in months:
        hub_id = f"HUB:{ym[0]}-{ym[1]:02d}"
        hub_id_of[ym] = hub_id
        nodes[hub_id] = Node(id=hub_id, label=month_label(*ym), kind="hub", column=2 * month_index[ym])

    for e in entries:
        if e.amount == 0:
            continue
        ym = (e.date.year, e.date.month)
        m_idx = month_index[ym]
        hub_id = hub_id_of[ym]
        node_id = unique_id(e.name, ym)
        if e.amount > 0:
            nodes[node_id] = Node(
                id=node_id, label=e.name, kind="source", column=2 * m_idx - 1, value=e.amount, date=e.date
            )
            links.append(Link(source=node_id, target=hub_id, value=e.amount, kind="income"))
        else:
            amt = abs(e.amount)
            nodes[node_id] = Node(
                id=node_id, label=e.name, kind="sink", column=2 * m_idx + 1, value=amt, date=e.date
            )
            links.append(Link(source=hub_id, target=node_id, value=amt, kind="expense"))

    running = 0.0
    for i, ym in enumerate(months):
        hub_id = hub_id_of[ym]
        income_total = sum(l.value for l in links if l.target == hub_id and l.kind == "income")
        expense_total = sum(l.value for l in links if l.source == hub_id and l.kind == "expense")
        total_in = running + income_total
        carry_out = total_in - expense_total
        nodes[hub_id].value = max(total_in, expense_total, abs(carry_out), 1e-9)
        nodes[hub_id].negative = carry_out < 0

        carry_kind = "carry_pos" if carry_out >= 0 else "carry_neg"
        if i + 1 < len(months):
            next_hub = hub_id_of[months[i + 1]]
            links.append(Link(source=hub_id, target=next_hub, value=abs(carry_out), kind=carry_kind))
        else:
            term_id = "TERMINAL"
            nodes[term_id] = Node(
                id=term_id,
                label=f"Final balance ({month_label(*ym)})",
                kind="terminal",
                column=nodes[hub_id].column + 1,
                value=abs(carry_out),
                negative=carry_out < 0,
            )
            links.append(Link(source=hub_id, target=term_id, value=abs(carry_out), kind=carry_kind))
        running = carry_out

    return SankeyGraph(nodes=list(nodes.values()), links=links)
