"""Shared FROM/AMOUNT/DATE workbook reader, used by both cashflowviz and cashflowviz.sankey.

Both visualizations read the same input shape — one row per cash movement,
positive AMOUNT for money in, negative for money out — and only differ in
how they lay it out.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass

import openpyxl


@dataclass(frozen=True)
class FlowEntry:
    name: str
    amount: float
    date: dt.date


def _to_date(value) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise ValueError(f"Cell does not contain a date: {value!r}")


def read_flow_entries(xlsx_path: str) -> list[FlowEntry]:
    """Read (FROM, AMOUNT, DATE) rows from the first sheet.

    The header row is located by name so column order/position doesn't
    matter. Entries are returned sorted chronologically (a stable sort, so
    same-day rows keep their original spreadsheet order).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row_idx = None
    col_of = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        names = {str(c.value).strip().lower(): c.column for c in row if c.value is not None}
        if {"from", "amount", "date"} <= names.keys():
            header_row_idx = row[0].row
            col_of = names
            break
    if header_row_idx is None:
        raise ValueError("Could not find header row with 'FROM', 'AMOUNT', 'DATE' columns")

    name_col = col_of["from"]
    amount_col = col_of["amount"]
    date_col = col_of["date"]

    entries: list[FlowEntry] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        name_cell = ws.cell(row=row[0].row, column=name_col).value
        amount_cell = ws.cell(row=row[0].row, column=amount_col).value
        date_cell = ws.cell(row=row[0].row, column=date_col).value
        if name_cell is None and amount_cell is None and date_cell is None:
            continue
        if name_cell is None or amount_cell is None or date_cell is None:
            continue
        entries.append(
            FlowEntry(name=str(name_cell).strip(), amount=float(amount_cell), date=_to_date(date_cell))
        )

    if not entries:
        raise ValueError("No data rows found in the workbook")

    entries.sort(key=lambda e: e.date)
    return entries
